# Guarda esto como:
# D:\Daniel\Desktop\Dashboard-active-inventory\excel_exporter.py

import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def get_cvss_rating(score_str):
    """Convierte una puntuación CVSS en texto (ej: '9.5' -> 'Crítica (Critical)')"""
    try:
        score = float(score_str)
        if score == 0.0:
            return "None"
        elif 0.1 <= score <= 3.9:
            return "Baja (Low)"
        elif 4.0 <= score <= 6.9:
            return "Media (Medium)"
        elif 7.0 <= score <= 8.9:
            return "Alta (High)"
        elif 9.0 <= score <= 10.0:
            return "Crítica (Critical)"
    except (ValueError, TypeError):
        return "Desconocida"
    return "Desconocida"

def replace_substring(test_str, s1, s2):
    """Función de ayuda para reemplazar texto."""
    if test_str is None:
        return ""
    result = ""
    i = 0
    while i < len(test_str):
        if test_str[i:i+len(s1)] == s1:
            result += s2
            i += len(s1)
        else:
            result += test_str[i]
            i += 1
    return result

def _generate_styled_excel(df, excel_file_name):
    """Función interna para crear y formatear el archivo Excel."""
    with ExcelWriter(excel_file_name, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Actives Inventory', index=False)

    wb = load_workbook(excel_file_name)
    ws = wb['Actives Inventory']

    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    cell_font = Font(name='Arial', size=12)
    cell_alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
    center_alignment = Alignment(vertical='center', horizontal='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar estilo a la cabecera
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Aplicar estilo a las celdas de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for i, cell in enumerate(row):
            if i == 0: # Centrar solo la primera columna (IP)
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment
            
            # Puedes cambiar este color de fondo si prefieres
            # cell.fill = PatternFill(start_color='eeeeee', end_color='eeeeee', fill_type='solid') 
            cell.font = cell_font
            cell.border = thin_border

    # Ajustar altura de filas con saltos de línea
    for row_idx in range(2, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and isinstance(cell_value, str):
                max_lines = max(max_lines, cell_value.count('\n') + 1)
        ws.row_dimensions[row_idx].height = 15 * max_lines


    # Definir anchos de columna
    widths = [25, 50, 120] # IP, Puertos, Vulnerabilidades
    for i, col_letter in enumerate(['A', 'B', 'C']):
        if i < len(widths):
            ws.column_dimensions[col_letter].width = widths[i]
    
    wb.save(excel_file_name)

def create_report(hosts_data, xlsx_file_name):
    """
    Función principal que la GUI llamará.
    Transforma los datos del dashboard y genera el Excel.
    """
    ips = []
    ports_services = []
    vulns = []

    for host_info in hosts_data:
        ports_servs = []
        vuls = []

        ips.append(host_info['ip'])
        
        # Formatear Puertos
        for port, service in host_info['ports']:
            ports_servs.append(f" - Port: {port}, Service: {service}")   
        ports_services.append("\n".join(ports_servs))

        # Formatear Vulnerabilidades
        # Ordenar por criticidad (opcional, pero recomendado)
        sorted_vulns = sorted(host_info['vulnerabilities'], key=lambda v: float(v['cvss']) if v['cvss'].replace('.', '', 1).isdigit() else 0.0, reverse=True)

        for vuln in sorted_vulns:
            formated_references = '\n\t'.join(vuln.get('references', []))
            
            cvss_score = vuln['cvss']
            cvss_rating = get_cvss_rating(cvss_score) # Usa la función de este archivo

            v =  f" - Port: {vuln['port']}, Service: {vuln['service']}\n"
            v += f"   Name: {vuln['name']}\n"
            v += f"   Criticidad: {cvss_score} ({cvss_rating})\n"
            v += f"   CVEs: {', '.join(vuln['cve'])}\n"
            v += f"   References: \n\t{formated_references}\n"
            # Puedes añadir descripción y fecha si quieres, como en tu script original
            # description = replace_substring(vuln.get('description', 'N/A'), "\n", "\n\t")
            # v += f"   Description:\n\t{description}\n"

            vuls.append(v)
        
        vulns.append("\n".join(vuls))

    datos = {
        "IP Direction": ips,
        "Ports and Services": ports_services,
        "Vulnerabilities": vulns
    }

    df = pd.DataFrame(datos)
    
    if not xlsx_file_name.endswith('.xlsx'):
        xlsx_file_name = f"{xlsx_file_name}.xlsx"

    try:
        _generate_styled_excel(df, xlsx_file_name)
        return True, None # (Éxito, Sin Error)
    except Exception as e:
        return False, str(e) # (Fallo, Mensaje de Error)